import tkinter as tk
from openpyxl import load_workbook, Workbook
from datetime import datetime
from enum import Enum

class Category(Enum):
    MISC = "Misc"
    APPLIANCE = "Appliance"
    BROADBAND = "Broadband"
    CAR_BIKE = "Car/ Bike"
    CASH_WITHDRAWAL = "Cash withdrawal"
    CLOTHES = "Clothes"
    COSMETICS_TOILETRY = "Cosmetics/ Toiletry"
    ELECTRONICS_GADGETS = "Electronics/ Gadgets"
    FESTIVAL = "Festival"
    FRUITS = "Fruits"
    FUEL = "Fuel"
    GIFT = "Gift"
    GOLD = "Gold"
    GROCERIES = "Groceries"
    HOUSEHOLD_ITEMS = "Household items"
    INSURANCE = "Insurance"
    INTERNET = "Internet"
    MEAT_EGGS_CHICKEN = "Meat/ Eggs/ Chicken"
    MEDICAL = "Medical"
    MOBILE_RECHARGE = "Mobile recharge"
    RESTAURANT = "Restaurant"
    SNACKS = "Snacks"
    SUBSCRIPTION = "Subscription"
    TAX = "Tax"
    TRAVEL = "Travel"
    VEGETABLES = "Vegetables"
    YOGA_FEES = "Yoga Fees"
    NURSERY = "Nursery"
    FLOWERS = "Flowers"
    SUPERMARKET = "SuperMarket"
    BOOKS = "Books"
    LAUNDRY = "Laundry"
    HOUSE_REPAIR = "House Repair"
    DRY_FRUITS = "Dry Fruits"
    SPORTS = "Sports"

class PaymentMode(Enum):
    GOOGLE_PAY = "Google Pay"
    PHONEPE = "PhonePe"
    SAMSUNG_PAY = "Samsung Pay"
    PNB_DEBIT_CARD = "PNB Debit Card"
    SBI_DEBIT_CARD = "SBI Debit Card"
    AXIS_BANK_DEBIT_CARD = "Axis Bank Debit Card"
    CASH = "Cash"
    UPI = "UPI"
    DEBIT_CARD = "Debit Card"
    CREDIT_CARD = "Credit Card"
    FUND_TRANSFER = "Fund Transfer"

def save_to_excel():
    date = date_entry.get()
    description = description_entry.get()
    category = category_var.get()
    payment_mode = payment_mode_var.get()
    entry_type = entry_type_entry.get()
    amount = amount_entry.get()
    month = month_entry.get()
    year = year_entry.get()

    filename = "Expense Tracker.xlsx"
    wb = None

    try:
        try:
            wb = load_workbook(filename)
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            headers = ["Date", "Description", "Category", "Payment mode", "Type", "Amount", "Month", "Year", "Month - Year"]
            sheet.append(headers)

        sheet = wb.active
        row_data = [date, description, category, payment_mode, entry_type, amount, month, year, f"{month}-{year}"]
        sheet.append(row_data)

        wb.save(filename)
        status_label.config(text=f"Data saved to {filename}")
    except Exception as e:
        status_label.config(text=f"Error: {e}")
    finally:
        if wb:
            wb.close()

root = tk.Tk()
root.minsize(400,400)
root.title("Expense Tracker Form")

# Create label
l = tk.Label(root, text="Expense Tracker 2.0")
l.config(font=("Courier", 14))
l.pack()

# Category dropdown
category_label = tk.Label(root, text="Category:")
category_label.pack()

category_var = tk.StringVar(root)
category_var.set(Category.MISC.value)  # Default value

category_dropdown = tk.OptionMenu(root, category_var, *[c.value for c in Category])
category_dropdown.pack()

# Payment mode dropdown
payment_mode_label = tk.Label(root, text="Payment Mode:")
payment_mode_label.pack()

payment_mode_var = tk.StringVar(root)
payment_mode_var.set(PaymentMode.GOOGLE_PAY.value)  # Default value

payment_mode_dropdown = tk.OptionMenu(root, payment_mode_var, *[p.value for p in PaymentMode])
payment_mode_dropdown.pack()

date_label = tk.Label(root, text="Date:")
date_label.pack()
date_entry = tk.Entry(root)
date_entry.pack()

description_label = tk.Label(root, text="Description:")
description_label.pack()
description_entry = tk.Entry(root)
description_entry.pack()

entry_type_label = tk.Label(root, text="Type:")
entry_type_label.pack()
entry_type_entry = tk.Entry(root)
entry_type_entry.pack()

amount_label = tk.Label(root, text="Amount:")
amount_label.pack()
amount_entry = tk.Entry(root)
amount_entry.pack()

month_label = tk.Label(root, text="Month:")
month_label.pack()
month_entry = tk.Entry(root)
month_entry.pack()

year_label = tk.Label(root, text="Year:")
year_label.pack()
year_entry = tk.Entry(root)
year_entry.pack()

save_button = tk.Button(root, text="Save to Excel", command=save_to_excel)
save_button.pack()

status_label = tk.Label(root, text="")
status_label.pack()

root.mainloop()
